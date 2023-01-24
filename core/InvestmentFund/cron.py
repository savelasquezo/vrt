import os

from openpyxl import Workbook
from openpyxl import load_workbook

from datetime import date


from django_cron import CronJobBase, Schedule
from django.db.models import F
from django.utils import timezone

from .models import Usuario


class AddFundsToUser(CronJobBase):
    RUN_EVERY_MINS = 1

    schedule = Schedule(run_every_mins=RUN_EVERY_MINS)
    code = 'InvestmentFund.AddFundsToUser'
    
    #crontab -e
    #*/1 * * * * /home/savelasquezo/savelasquezo/VaorTrading/venv/bin/python /home/savelasquezo/savelasquezo/VaorTrading/core/manage.py runcrons
    def do(self):
        
        InfoUser = Usuario.objects.all()
        Usuario.objects.filter(date_expire__lte=timezone.now()).update(is_operating=False)

        for nUser in InfoUser:

            if nUser.is_operating:

                CUser = Usuario.objects.filter(id=nUser.id)
                    
                if date.today().day == 1:
                    CUser.update(available_tickets=3)

                cAmmount = int(nUser.ammount)
                cTotalInterest = int(nUser.total_interest)
                cPaid = int(nUser.paid)
                cPaidRef = int(nUser.ref_paid)

                cInterest = float((nUser.interest)/(100*30))
                cInterestRef = float((nUser.ref_interest)/(100*30))

                cValue = int(cAmmount*(cInterest))
                cValueRef = int(cAmmount*(cInterestRef))
                
                cAvailable = int(cTotalInterest-cPaid + cValue)

                CUser.update(
                    available=cAvailable,
                    total_interest=F('total_interest') + cValue)

                if nUser.ref_id:
                    cValueRef = int(cAmmount*(cInterestRef))
                    CUser.update(ref_total=F('ref_total') + cValueRef)
                
                UserRef = Usuario.objects.filter(ref_id= nUser.codigo)
                mAviableUserRef = 0
                
                for mUser in UserRef:
                    mAviableUserRef += int(mUser.ref_total)

                mAviableUserTotal = mAviableUserRef - cPaidRef
                cTodayRef =  mAviableUserTotal - nUser.ref_available
                
                CUser.update(
                    ref_available=mAviableUserTotal,
                    total_ref=mAviableUserRef
                    )

                CUser.update(total=F('total_ref') + F('total_interest'))

                FileName = '/home/savelasquezo/apps/vrt/core/logs/users/'+ nUser.username + '.xlsx'


                if not os.path.exists(FileName):
                    WB = Workbook()
                    WS = WB.active
                    WS.append(["Tipo","Fecha", "$Interes", "$Comiciones", "AcInteres", "AcComisiones", "$Ticket", "Origen", "Total"])
                else:
                    WB = load_workbook(FileName)
                    WS = WB.active
                
                cTotal = nUser.total
                cAviableRef = nUser.ref_available
                NowToday = timezone.now().strftime("%Y-%m-%d %H:%M")

                FileData = [1, NowToday, cValue, cTodayRef, cAvailable, cAviableRef, "", "", cTotal]

                WS.append(FileData)
                WB.save(FileName)

cronjobs = [
    AddFundsToUser,
]